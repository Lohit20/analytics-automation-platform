"""Write pipeline outputs: Power-BI-ready CSVs and a formatted Excel business report."""

import pandas as pd
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

import config


def write_csvs(history_df: pd.DataFrame, kpi_df: pd.DataFrame) -> None:
    """Write tidy CSVs that Power BI (or any BI tool) can connect to directly."""
    history_df.to_csv(config.CSV_RATES_PATH, index=False)
    kpi_df.to_csv(config.CSV_KPI_PATH, index=False)


def _autosize_and_style_header(ws) -> None:
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for col_cells in ws.columns:
        length = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(12, length + 2)


def write_excel_report(history_df: pd.DataFrame, kpi_df: pd.DataFrame) -> None:
    """Build a multi-sheet Excel workbook: KPI summary, raw history, and a trend chart."""
    with pd.ExcelWriter(config.EXCEL_REPORT_PATH, engine="openpyxl") as writer:
        kpi_df.to_excel(writer, sheet_name="KPI Summary", index=False)
        history_df.to_excel(writer, sheet_name="Rate History", index=False)

        _autosize_and_style_header(writer.sheets["KPI Summary"])
        _autosize_and_style_header(writer.sheets["Rate History"])

    _add_trend_chart(history_df)


def _add_trend_chart(history_df: pd.DataFrame) -> None:
    """Add a line chart of each currency's rate over time to its own sheet."""
    from openpyxl import load_workbook

    wb = load_workbook(config.EXCEL_REPORT_PATH)
    pivot = history_df.pivot(index="date", columns="currency", values="rate").reset_index()

    ws = wb.create_sheet("Rate Trend Chart")
    ws.append(["date"] + list(pivot.columns[1:]))
    for row in pivot.itertuples(index=False):
        ws.append([row.date.strftime("%Y-%m-%d")] + list(row[1:]))

    chart = LineChart()
    chart.title = "FX Rate Trend by Currency"
    chart.y_axis.title = f"Rate (per 1 {config.BASE_CURRENCY})"
    chart.x_axis.title = "Date"

    n_rows = ws.max_row
    n_cols = ws.max_column
    data = Reference(ws, min_col=2, max_col=n_cols, min_row=1, max_row=n_rows)
    cats = Reference(ws, min_col=1, min_row=2, max_row=n_rows)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.width, chart.height = 24, 12

    ws.add_chart(chart, "H2")
    wb.save(config.EXCEL_REPORT_PATH)
